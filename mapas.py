import openpyxl
import sys

try:
    file_path = sys.argv[1]
    map_name = sys.argv[2]
    equipamento = sys.argv[3]
    folha = sys.argv[4]
except:
    print(f'''
    Ta faltando argumento amigão

    {sys.argv[0]} path_planilha.xlms map.txt equipamento nome_da_pagina

    ''')
    sys.exit(1)
    

def saveFile(content, file_name):
    print(content, file_name)
    with open(f'./map_files/mapa{file_name}.txt', 'w') as file:
        file.write(content)

wb = openpyxl.load_workbook(file_path)
sheet = wb[folha]

sh = wb.active

for rowOfCellObjects in sheet['A1':f'A{sh.max_row}']:
    for cellObj in rowOfCellObjects:
        with open(map_name, 'r') as file:
            try:
                nameFile = cellObj.value.split('_')
            except:
                pass
            data = file.read()
            try:
                data = data.replace(f'{equipamento}', cellObj.value)
            except:
                pass
        saveFile(data, nameFile[0])

print(f"""
==============================================
    Linhas totais geradas do arquivo :)
    {sh.max_row}
==============================================
""")